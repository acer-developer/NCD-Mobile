#!/usr/bin/env python3
"""
build_enriched.py  -  NCD Mobile enrichment workbook builder

Reads the original "NCD Mobile.xlsx" plus the accumulated research in
contacts.json, and writes a color-coded "NCD Mobile - Enriched.xlsx".

Every ADDED contact cell is filled by surety/confidence:
    high   -> GREEN   (verified: official site / annual report / MCA / offer doc, or multi-source)
    medium -> YELLOW  (plausible: single source, LinkedIn, inferred email pattern)
    low    -> RED     (weak / fallback: generic landline used, unverified guess)

Nothing is invented here - this script only paints what research put into
contacts.json. The routine (see README.md) is what fills contacts.json.

Run any time:  python build_enriched.py
It is idempotent - safe to run after every batch.
"""
import json
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HERE = Path(__file__).parent
ORIG = HERE / "NCD Mobile.xlsx"
CONTACTS = HERE / "contacts.json"
OUT = HERE / "NCD Mobile - Enriched.xlsx"

# confidence -> fill color (light shades so text stays readable)
FILLS = {
    "high":   PatternFill("solid", fgColor="C6EFCE"),   # green
    "medium": PatternFill("solid", fgColor="FFEB9C"),   # yellow
    "low":    PatternFill("solid", fgColor="FFC7CE"),   # red
}
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BASE_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# new columns appended after the original ones
NEW_COLS = [
    "Contact Name",
    "Designation",
    "Direct Mobile",
    "Direct Email",
    "Fallback (Landline/IR)",
    "Confidence",
    "Source",
    "Notes",
    "Run",
]


def main():
    data = json.loads(CONTACTS.read_text(encoding="utf-8"))
    companies = {c["row"]: c for c in data["companies"]}

    df = pd.read_excel(ORIG)
    df.to_excel(OUT, index=False)  # start from an exact copy of the original

    wb = load_workbook(OUT)
    ws = wb.active

    n_orig = df.shape[1]
    # write new headers
    for j, name in enumerate(NEW_COLS):
        c = ws.cell(row=1, column=n_orig + 1 + j, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    # style original header row too
    for j in range(1, n_orig + 1):
        c = ws.cell(row=1, column=j)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)

    filled = {"high": 0, "medium": 0, "low": 0}
    done = 0
    for i in range(len(df)):
        rec = companies.get(i)
        xlrow = i + 2  # header is row 1
        if not rec:
            continue
        ct = rec.get("contact", {})
        conf = (ct.get("confidence") or "").lower()
        fill = FILLS.get(conf)
        vals = [
            ct.get("name"),
            ct.get("designation"),
            ct.get("mobile"),
            ct.get("email"),
            ct.get("fallback"),
            (conf.upper() if conf else None),
            ct.get("source"),
            ct.get("notes"),
            ct.get("run"),
        ]
        any_val = any(v for v in vals)
        for j, v in enumerate(vals):
            cell = ws.cell(row=xlrow, column=n_orig + 1 + j, value=v)
            cell.font = BASE_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill and any_val:
                cell.fill = fill
        if any_val and conf in filled:
            filled[conf] += 1
        if rec.get("status") == "done":
            done += 1

    # widths
    widths = [34, 14, 8, 55, 22, 12, 22, 55, 40, 20, 20, 55, 22, 26, 26, 18, 24]
    from openpyxl.utils import get_column_letter
    for idx in range(1, ws.max_column + 1):
        w = widths[idx - 1] if idx - 1 < len(widths) else 20
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"

    # legend sheet
    lg = wb.create_sheet("Legend")
    rows = [
        ("NCD Mobile - Enrichment legend", ""),
        ("", ""),
        ("Color", "Meaning (surety of the added contact)"),
        ("GREEN", "High - verified: official site / annual report / MCA / NCD offer doc, or 2+ sources agree"),
        ("YELLOW", "Medium - plausible: single source, LinkedIn, or inferred email pattern"),
        ("RED", "Low - weak / fallback: generic landline used instead of a person, or unverified"),
        ("", ""),
        ("Priority order", "CFO -> Treasury head -> Finance team -> Fallback (landline / IR desk)"),
        ("Rule", "Never fabricate a number. If nothing found, leave blank and mark status not_found."),
    ]
    for r_i, (a, b) in enumerate(rows, start=1):
        lg.cell(row=r_i, column=1, value=a).font = Font(name="Arial", bold=(r_i in (1, 3)), size=11)
        lg.cell(row=r_i, column=2, value=b).font = Font(name="Arial", size=10)
    for k, col in (("GREEN", 4), ("YELLOW", 5), ("RED", 6)):
        lg.cell(row=col, column=1).fill = FILLS[{"GREEN": "high", "YELLOW": "medium", "RED": "low"}[k]]
    lg.column_dimensions["A"].width = 16
    lg.column_dimensions["B"].width = 95

    wb.save(OUT)
    print(f"Wrote {OUT.name}")
    print(f"  companies done:   {done}/{len(df)}")
    print(f"  green (high):     {filled['high']}")
    print(f"  yellow (medium):  {filled['medium']}")
    print(f"  red (low):        {filled['low']}")


if __name__ == "__main__":
    main()
